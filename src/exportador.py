import pandas as pd
import os
import requests
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
import subprocess
import platform

try:
    from config import APPS_SCRIPT_URL
except ImportError:
    APPS_SCRIPT_URL = ''

def obtener_calificaciones_sheets():
    if not APPS_SCRIPT_URL or APPS_SCRIPT_URL.startswith('PEGA'):
        print("⚠ APPS_SCRIPT_URL no configurada en exportador.py — calificaciones no se incluirán")
        return {}
    try:
        res = requests.get(f"{APPS_SCRIPT_URL}?action=getAll", timeout=10)
        data = res.json()
        if data.get('ok'):
            cal = data.get('data', {})
            print(f"✓ {len(cal)} calificaciones obtenidas desde Google Sheets")
            return cal
        else:
            print(f"⚠ Google Sheets devolvió error: {data}")
            return {}
    except Exception as e:
        print(f"⚠ No se pudo conectar a Google Sheets: {e}")
        return {}


class Exportador:
    def cerrar_excel(self, nombre_archivo):
        try:
            sistema = platform.system()
            
            if sistema == "Windows":
                subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], 
                             capture_output=True, stderr=subprocess.DEVNULL)
                print("✓ Instancias de Excel cerradas")
            elif sistema == "Linux":
                subprocess.run(['pkill', '-9', 'libreoffice'], 
                             capture_output=True, stderr=subprocess.DEVNULL)
                subprocess.run(['pkill', '-9', 'soffice'], 
                             capture_output=True, stderr=subprocess.DEVNULL)
                print("✓ Instancias de LibreOffice cerradas")
        except Exception as e:
            pass
    
    def exportar_excel(self, datos, nombre_archivo=None):
        if not datos:
            print("\n⚠ No hay datos para exportar")
            return None
        
        if nombre_archivo is None:
            now = datetime.now()
            nombre_archivo = now.strftime("%m.%d.%Y") + "-T2MDC.xlsx"
        
        output_dir = "resultados_seace"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        ruta_completa = os.path.join(output_dir, nombre_archivo)
        
        self.cerrar_excel(ruta_completa)

        # ── Obtener calificaciones guardadas en Google Sheets ──────────
        calificaciones = obtener_calificaciones_sheets()
        
        df_raw = pd.DataFrame(datos)
        
        fecha_extraccion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        df_procesado = pd.DataFrame()
        
        df_procesado['Item'] = range(1, len(df_raw) + 1)
        df_procesado['Fecha de Extracción'] = fecha_extraccion
        df_procesado['Fecha y Hora de Publicacion'] = df_raw.get('Fecha y Hora de Publicacion', '')
        df_procesado['Nro de Proceso'] = df_raw.get('Nomenclatura', '')
        df_procesado['Entidad (Empresa)'] = df_raw.get('Nombre o Sigla de la Entidad', '')
        df_procesado['Tipo Servicio'] = df_raw.get('Objeto de Contratación', '')
        
        df_procesado['Registro de Participantes (Fecha fin)'] = ''
        df_procesado['Hora de Envío'] = ''
        df_procesado['Fecha de Formulacion de consultas'] = ''
        df_procesado['Fecha de integracion de bases'] = ''
        df_procesado['Presentacion de Propuesta'] = ''
        df_procesado['Hora de Envío.1'] = ''
        
        for i, dato in df_raw.iterrows():
            cronograma = dato.get('Cronograma', [])
            if isinstance(cronograma, list):
                for etapa in cronograma:
                    etapa_nombre = etapa.get('Etapa', '')
                    if 'Registro de participantes' in etapa_nombre or 'Invitación' in etapa_nombre:
                        df_procesado.at[i, 'Registro de Participantes (Fecha fin)'] = etapa.get('Fecha Fin', '')
                    elif 'Formulación de consultas' in etapa_nombre:
                        df_procesado.at[i, 'Fecha de Formulacion de consultas'] = etapa.get('Fecha Fin', '')
                    elif 'Integración de las Bases' in etapa_nombre:
                        df_procesado.at[i, 'Fecha de integracion de bases'] = etapa.get('Fecha Fin', '')
                    elif 'Presentación de propuestas' in etapa_nombre:
                        df_procesado.at[i, 'Presentacion de Propuesta'] = etapa.get('Fecha Fin', '')
        
        df_procesado['Descripción del RQ'] = df_raw.get('Descripción de Objeto', '')
        
        df_procesado['Estado'] = df_raw.get('Estado', '-')
        df_procesado['Duración (días)'] = df_raw.get('Duracion_dias', '-')
        df_procesado['Ciudad'] = df_raw.get('Ciudad', '-')
        df_procesado['Modalidad de trabajo'] = df_raw.get('Modalidad_trabajo', '-')
        df_procesado['Experiencia Requerida'] = df_raw.get('Experiencia_requerida', '-')
        df_procesado['Experiencia Minima (MYPE)'] = df_raw.get('Experiencia_minima_mype', '-')
        df_procesado['Servicios similares válidos'] = df_raw.get('Servicios_similares', '-')
        
        df_procesado['Documento'] = df_raw.get('Documento_URL', '')

        # ── Merge columna Calificacion desde Google Sheets ─────────────
        df_procesado['Calificacion'] = df_procesado['Nro de Proceso'].apply(
            lambda nro: calificaciones.get(str(nro).strip(), 'NO')
        )
        
        def convertir_fecha(fecha_str):
            try:
                return pd.to_datetime(fecha_str, format='%d/%m/%Y %H:%M')
            except:
                return pd.NaT
        
        df_procesado['_fecha_temp'] = df_procesado['Fecha y Hora de Publicacion'].apply(convertir_fecha)
        df_procesado = df_procesado.sort_values('_fecha_temp', ascending=False, na_position='last')
        df_procesado = df_procesado.drop('_fecha_temp', axis=1)
        df_procesado = df_procesado.reset_index(drop=True)
        df_procesado['Item'] = range(1, len(df_procesado) + 1)
        
        with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:
            df_procesado.to_excel(writer, sheet_name='Procedimientos 2026', index=False)
            worksheet = writer.sheets['Procedimientos 2026']
            
            columnas_anchos = {
                'A': 8,   'B': 22,  'C': 22,  'D': 35,  'E': 40,  'F': 18,  
                'G': 25,  'H': 15,  'I': 25,  'J': 25,  'K': 25,  'L': 15,  
                'M': 60,  'N': 15,  'O': 18,  'P': 20,  'Q': 25,  'R': 25,  
                'S': 25,  'T': 60,  'U': 60,  'V': 15
            }
            
            for col, ancho in columnas_anchos.items():
                worksheet.column_dimensions[col].width = ancho
            
            header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            header_font = Font(bold=True)
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            # Colorear columna Calificacion (V) según valor
            cal_fill_si = PatternFill(start_color="14532d", end_color="14532d", fill_type="solid")
            cal_fill_no = PatternFill(start_color="3f3f46", end_color="3f3f46", fill_type="solid")

            for row in range(2, len(df_procesado) + 2):
                # Documento URL (columna U)
                cell_doc = worksheet[f'U{row}']
                doc_url = cell_doc.value
                if doc_url and doc_url.strip():
                    cell_doc.font = Font(color="0563C1", underline="single")
                    cell_doc.alignment = Alignment(horizontal="center", vertical="center")

                # Calificacion (columna V)
                cell_cal = worksheet[f'V{row}']
                cell_cal.alignment = Alignment(horizontal="center", vertical="center")
                if str(cell_cal.value).strip() == 'SI':
                    cell_cal.fill = cal_fill_si
                    cell_cal.font = Font(color="4ade80", bold=True)
                else:
                    cell_cal.fill = cal_fill_no
                    cell_cal.font = Font(color="a1a1aa")
        
        print(f"\n{'='*60}")
        print(f"✓ EXPORTADO EXITOSAMENTE")
        print(f"{'='*60}")
        print(f"Archivo: {ruta_completa}")
        print(f"Total registros: {len(datos)}")
        cal_si = sum(1 for v in calificaciones.values() if v == 'SI')
        print(f"Calificaciones SI aplicadas: {cal_si}")
        print(f"Ordenado por fecha (más reciente primero)")
        print(f"{'='*60}\n")
        
        return os.path.abspath(ruta_completa)